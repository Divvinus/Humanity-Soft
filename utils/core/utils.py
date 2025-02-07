import asyncio
import random
import openpyxl

from typing import Optional, Literal, Callable, Any, Awaitable
from functools import wraps
from undetected_playwright.async_api import Locator, Error

from .logger import logger


StateType = Literal["visible", "hidden", "attached", "detached"]

class ElementHandlingError(Exception):
    pass


def with_retry(locator_param_name: str = 'locator'):
    def decorator(func: Callable) -> Callable:
        @wraps(func)
        async def wrapper(self, *args: Any, **kwargs: Any) -> Optional[Any]:
            max_retries = kwargs.pop('max_retries', 3)
            retry_delay = kwargs.pop('retry_delay', 1)
            
            locator = kwargs.get(locator_param_name) or (args[0] if args else None)
            
            for attempt in range(max_retries):
                try:
                    return await func(self, *args, **kwargs)
                
                except Exception as e:
                    if attempt == max_retries - 1:
                        logger.error(f"Элемент: {locator} | Превышено максимальное количество попыток: {str(e)}")
                        raise
                    
                    logger.warning(f"Элемент: {locator} | Попытка {attempt + 1}/{max_retries} не удалась: {str(e)}")
                    await asyncio.sleep(retry_delay)
            
            return None
        return wrapper
    return decorator

class ElementHandler:
    @with_retry(locator_param_name='locator')
    async def handle_element(
        self,
        locator: Locator,
        state: StateType = "visible",
        timeout: int = 30_000,
        max_retries: int = 3,
        retry_delay: int = 1,
        custom_action: Optional[Callable[[Locator], Awaitable[Any]]] = None
    ) -> bool:
        try:
            await locator.wait_for(state=state, timeout=timeout)

            if not await locator.is_enabled():
                raise ElementHandlingError("Элемент недоступен для взаимодействия")

            if custom_action:
                await custom_action(locator)
            else:
                await locator.hover()
                await locator.click(force=True, timeout=5000, no_wait_after=True)
                
            return True
            
        except TimeoutError as e:
            logger.warning(f"Таймаут ожидания элемента {locator} в состоянии {state}")
            raise ElementHandlingError(f"Timeout: {str(e)}")
            
        except Error as e:
            logger.error(f"Ошибка Playwright: {str(e)}")
            raise ElementHandlingError(f"Playwright error: {str(e)}")
            
        except Exception as e:
            logger.error(f"Непредвиденная ошибка: {str(e)}")
            raise ElementHandlingError(f"Unexpected error: {str(e)}")
    
async def smart_sleep(up, to):
    duration = random.randint(up, to)
    logger.info(f"💤 Ожидаем {duration:.2f} секунд")
    await asyncio.sleep(duration)

def get_accounts_data():
    try:
        book = openpyxl.load_workbook('data/accounts.xlsx', read_only=True)
        sheet = book.active
    except FileNotFoundError:
        logger.error("Файл accounts.xlsx не найден")
        return [], [], []
    except Exception as e:
        logger.error(f"Ошибка при открытии файла: {e}")
        return [], [], []

    accounts = []
    
    for row in range(2, sheet.max_row + 1):
        try:
            account_name = sheet.cell(row=row, column=1).value
            private_key = sheet.cell(row=row, column=2).value
            proxies = sheet.cell(row=row, column=3).value

            if not all([account_name, private_key]):
                continue

            accounts.append({
                'account_name': account_name,
                'private_key': private_key,
                'proxies': proxies
            })
        except Exception as e:
            logger.error(f"Ошибка при чтении строки {row}: {e}")

    return accounts
